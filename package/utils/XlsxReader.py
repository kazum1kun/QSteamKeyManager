import re

from package.ENV import ENV
from openpyxl import load_workbook

class XlsxReader:
    """Support for MS Excel/LibreOffice Calc file format"""

    @staticmethod
    def read(input_file):
        """Parse the input file row by row and try to extract kay/game/notes info"""
        wb = load_workbook(input_file)
        ws = wb.active

        # Read a few rows and determine if there exists a column where keys can be found
        for i in range (1, ws.max_row):
            # Skip empty rows
            if len(ws[i]) == 0:
                continue

            # Attempt matching
            if